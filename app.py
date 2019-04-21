import os
import tabula
from flask import Flask, render_template, request
from os.path import join, dirname, realpath
from openpyxl import load_workbook

from werkzeug.utils import secure_filename

app = Flask(__name__)

UPLOAD_FOLDER = join(dirname(realpath(__file__)), 'static/uploads/')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

result_uploaded = False
data = {}


@app.route('/', methods=["GET", "POST"])
def index():

    global result_uploaded
    global data

    if request.method == "GET":
        return render_template("index.html")

    else:
        if data == {}:
            file = request.files['file']
            file.filename = "results.pdf"
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

            data = process_pdf()
            alter_data_for_long_names()
            return render_template("index.html", pdf_uploaded=True)

        else:
            file = request.files['file']
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            write_excel(filename)
            return render_template("index.html", excel_processed=True, filename=filename)


def process_pdf():
    data_table = tabula.read_pdf(str(UPLOAD_FOLDER + "/results.pdf"), multiple_tables=True, pages="all")
    names = {}
    subs = []
    for iter_1 in range(len(data_table)):
        skip = 0
        if "Subject Code" in data_table[iter_1][1][0]:
            subs = []
            for iter_2 in range(len(data_table[iter_1].columns)):
                subs.append(data_table[iter_1][iter_2][0])
            skip = 2

        for iter_2 in range(skip, len(data_table[iter_1])):
            if type(data_table[iter_1][0][iter_2]) == str:
                roll_no = str(data_table[iter_1][0][iter_2].split()[0]).strip()

            if roll_no not in names:
                names[roll_no] = []

            name = {}
            skip = 0
            if (len(data_table[iter_1].columns) - len(subs) + 2) < 2:
                skip = 1
            for iter_3 in range(2, len(data_table[iter_1].columns)):
                name[subs[iter_3 + skip]] = data_table[iter_1][iter_3][iter_2]

            names[roll_no].append(name)

    return names


def alter_data_for_long_names():
    global data
    for name in data:
        if len(data[name]) > 1:
            if data[name][-1].keys() == data[name][-2].keys():
                data[name].pop(-1)


def write_excel(file):
    global data
    book = load_workbook(UPLOAD_FOLDER + file)
    sheets = book.get_sheet_names()
    filters = ["IV", "III", "II", "I"]

    sheet_names = []
    for sheet in sheets:
        if sheet.split()[0] in filters and sheet.split()[1] != "SEM":
            sheet_names.append(sheet)
    print(data)

    missing_roll_nos = []

    for sheet_name in sheet_names:
        sheet = book.get_sheet_by_name(sheet_name)
        subs = []

        for iter_1 in range(5, 15):
            if sheet.cell(row=8, column=iter_1).value is None:
                break
            else:
                subs.append(str(sheet.cell(row=8, column=iter_1).value).split()[0])

        for iter_1 in range(10, 200):
            if sheet.cell(row=iter_1, column=2).value is None:
                break
            else:
                roll_no = str(sheet.cell(row=iter_1, column=2).value).strip()
                print("ROLL - NO" + roll_no)
                for iter_2 in range(len(subs)):
                    try:
                        if str(data[roll_no][-1][subs[iter_2]]) != "nan":
                            sheet.cell(row=iter_1, column=iter_2 + 5, value = (data[roll_no][-1][subs[iter_2]] if (len(data[roll_no][-1][subs[iter_2]]) > 0 and data[roll_no][-1][subs[iter_2]] != "NC") else str(sheet.cell(row=iter_1, column=iter_2 + 5).value)))
                    except KeyError:
                        missing_roll_nos.append(roll_no)

        book.save(UPLOAD_FOLDER + file)

    for roll_no in set(missing_roll_nos):
        print(roll_no)

    book.save(UPLOAD_FOLDER + file)


if __name__ == "__main__":
    app.secret_key = os.urandom(12)
    app.run(host='0.0.0.0', debug=True, threaded=True)
